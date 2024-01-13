# -*- coding: utf-8 -*-

import copy
from openpyxl.utils import get_column_letter
from openpyxl.cell.text import InlineFont
from .cellcontext import CellContextX
from openpyxl.utils.cell import range_boundaries

class SheetBase():

    def copy_sheet_settings(self):
        self.wtsheet.sheet_format = copy.copy(self.rdsheet.sheet_format)
        self.wtsheet.sheet_properties = copy.copy(self.rdsheet.sheet_properties)
        # copy print settings
        self.wtsheet.page_setup = copy.copy(self.rdsheet.page_setup)
        self.wtsheet.print_options = copy.copy(self.rdsheet.print_options)
        self.wtsheet._print_rows = copy.copy(self.rdsheet._print_rows)
        self.wtsheet._print_cols = copy.copy(self.rdsheet._print_cols)
        self.wtsheet._print_area = copy.copy(self.rdsheet._print_area)
        self.wtsheet.page_margins = copy.copy(self.rdsheet.page_margins)
        self.wtsheet.protection = copy.copy(self.rdsheet.protection)
        self.wtsheet.HeaderFooter = copy.copy(self.rdsheet.HeaderFooter)
        self.wtsheet.views = copy.copy(self.rdsheet.views)
        self.wtsheet._images = copy.copy(self.rdsheet._images)
        # copy of tables doesn't work, add them individually
        for table in self.rdsheet.tables:
            self.wtsheet.add_table(self.rdsheet.tables[table])

    def setup_tables(self, sheet):
        return {table_name: Table(sheet.tables[table_name].ref, table_name) for table_name in sheet.tables}          
        
    def copy_row_dimension(self, rdrowx, wtrowx):
        if wtrowx in self.wtrows:
            return
        dim = self.rdsheet.row_dimensions.get(rdrowx)
        if dim:
            self.wtsheet.row_dimensions[wtrowx] = copy.copy(dim)
            self.wtsheet.row_dimensions[wtrowx].worksheet = self.wtsheet
            self.wtrows.add(wtrowx)

    def copy_col_dimension(self, rdcolx, wtcolx):
        if wtcolx in self.wtcols:
            return
        rdkey = get_column_letter(rdcolx)
        rddim = self.rdsheet.column_dimensions.get(rdkey)
        if not rddim:
            return
        wtdim = copy.copy(rddim)
        if rdcolx != wtcolx:
            wtkey = get_column_letter(wtcolx)
            wtdim.index = wtkey
            d = wtcolx - rdcolx
            wtdim.min += d
            wtdim.max += d
        else:
            wtkey = rdkey
        self.wtsheet.column_dimensions[wtkey] = wtdim
        self.wtsheet.column_dimensions[wtkey].worksheet = self.wtsheet
        self.wtcols.add(wtcolx)

    def _cell(self, source_cell, rdrowx, rdcolx, wtrowx, wtcolx, value=None, data_type=None):
        target_cell = self.wtsheet.cell(column=wtcolx, row=wtrowx)
        if value is None:
            target_cell.value = source_cell._value
            target_cell.data_type = source_cell.data_type
        elif isinstance(value, STRING_TYPES) and value.startswith('='):
            target_cell.value = value
        elif data_type:
            target_cell._value = value
            target_cell.data_type = data_type
        else:
            #value, data_type = get_type(value)
            target_cell.value = value
            #target_cell.data_type = data_type
        if source_cell.has_style:
            target_cell._style = copy.copy(source_cell._style)
        if source_cell.hyperlink:
            target_cell._hyperlink = copy.copy(source_cell.hyperlink)
        #if source_cell.comment:
        #    target_cell.comment = copy.copy(source_cell.comment)
        return target_cell

    def cell(self, source_cell, rdrowx, rdcolx, wtrowx, wtcolx, value=None, data_type=None):
        self.copy_row_dimension(rdrowx, wtrowx)
        self.copy_col_dimension(rdcolx, wtcolx)
        return self._cell(source_cell, rdrowx, rdcolx, wtrowx, wtcolx, value, data_type)

    def get_cell_context(self, cell_node, rv, cty):
        return CellContextX(self, cell_node, rv, cty)


class BookBase():

    def get_font(self, fontId):
        ifont = self.font_map.get(fontId)
        if ifont:
            return ifont
        else:
            font = self.workbook._fonts[fontId]
            ifont = InlineFont()
            ifont.rFont = font.name
            ifont.charset = font.charset
            ifont.family = font.family
            ifont.b = font.b
            ifont.i = font.i
            ifont.strike = font.strike
            ifont.outline = font.outline
            ifont.shadow = font.shadow
            ifont.condense = font.condense
            ifont.extend = font.extend
            ifont.color = font.color
            ifont.sz = font.sz
            ifont.u = font.u
            ifont.vertAlign = font.vertAlign
            ifont.scheme = font.scheme
            self.font_map[fontId] = ifont
            return ifont


class Table(object):
    '''
    assume tables only expand downwards, i.e. more rows
    '''
    
    def __init__(self, ref, name):
        self.min_col, self.min_row, self.max_col, self.max_row = range_boundaries(ref)
        self.name = name 
        
    def __repr__(self):
        class_name = type(self).__name__
        return (f'{class_name}, table name:{self.name}\n'
                f'..min_row: {self.min_row}, min_col:{self.min_col}\n'
                f'..max_row: {self.max_row}, max_col:{self.max_col}\n'
                )
            
    def reset(self, sheet_table):
        '''
        resets the position and size of the actual table in the worksheet 
        '''
        top_left = get_column_letter(self.min_col) + str(self.min_row)
        bottom_right = get_column_letter(self.max_col) + str(self.max_row)
        sheet_table.ref = f'{top_left}:{bottom_right}'
        
    def reset_pos(self, cell_rowx, cell_colx):
        ''' 
        start of new table in output - reset position of top left and  size
        '''
        self.min_row = cell_rowx
        self.min_col = cell_colx
        self.max_row = self.min_row
        self.max_col = self.min_col
        
    def expand(self, cell_rowx, cell_colx):
        self.max_row = max(self.max_row, cell_rowx)
        self.max_col = max(self.max_col, cell_colx)
        
    def is_cell_inside(self, cell_rowx, cell_colx): 
        ''' 
        test if cell is inside this table, used to identify if table needs to expand
        '''
        return ((self.min_row <= cell_rowx <= self.max_row)
            and (self.min_col <= cell_colx <= self.max_col) 
            )
